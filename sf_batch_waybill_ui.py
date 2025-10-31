"""批量顺丰运单 PDF 生成工具

功能 (对应需求步骤):
1. Tkinter UI 顶部右上角置顶: 
   - 第一行: [选择Excel] 按钮 -> 选择含有 "序号" 与 "物流单号" 列的 Excel 文件 (D)
   - 第二行: 文本输入框 (输入数字序号) + 按钮 [序号] -> 保存为 order_no
   - 第三行: 按钮 [确认] [下一单] [结束]
2. Excel 中查找列名 "序号" 和 "物流单号" (区分大小写, 去除首尾空格再匹配)
3. 在 "序号" 列查找值 == order_no 的单元格行号 -> row_now (内部用 0-based 索引, 展示给用户 1-based)
4. 提取该行 "物流单号" -> current_order_no
5. 打开 https://www.sf-express.com/chn/sc/waybill/waybill-detail/<current_order_no>
6. 等待页面打开(用户可手动等待验证码弹出并输入)
7. 用户手动点击展开详情
8. 用户在 UI 中点击 [确认]
9. 生成 PDF 保存到 output/<运单号>.pdf
10. 用户点击 [下一单]
11. row_now += 1
12. 读取新行物流单号, 若= "END" 则程序结束; 否则重复 5~11

注意: 不做自动验证码 / 展开详情; 不做列名模糊匹配; Excel 文件在首次成功选择后可复用。
"""
from __future__ import annotations
import os
import sys
import base64
import threading
from dataclasses import dataclass
from typing import Optional, List

import tkinter as tk
from tkinter import filedialog, messagebox

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    print("缺少 openpyxl 库: 请先运行 'pip install openpyxl' 再启动程序或重新打包。")
    import sys as _sys
    raise SystemExit(1)
try:
    from selenium import webdriver
    from selenium.webdriver.edge.options import Options as EdgeOptions
    from selenium.webdriver.remote.webdriver import WebDriver
except ModuleNotFoundError:
    print("缺少 selenium 库: 请先运行 'pip install selenium' 再启动程序或重新打包。")
    import sys as _sys
    raise SystemExit(1)

BASE_URL = "https://www.sf-express.com/chn/sc/waybill/waybill-detail/{waybill}"


def _detect_edge_binary() -> Optional[str]:
    candidates = [
        r"C:\\Program Files\\Microsoft\\Edge\\Application\\msedge.exe",
        r"C:\\Program Files (x86)\\Microsoft\\Edge\\Application\\msedge.exe",
    ]
    for p in candidates:
        if os.path.exists(p):
            return p
    return None


def create_driver(headless: bool = False) -> WebDriver:
    opts = EdgeOptions()
    if headless:
        opts.add_argument("--headless=new")
    opts.add_argument("--disable-gpu")
    opts.add_argument("--window-size=1280,900")
    binary = _detect_edge_binary()
    if binary:
        opts.binary_location = binary
    driver = webdriver.Edge(options=opts)
    driver.set_page_load_timeout(60)
    return driver


def print_to_pdf(driver: WebDriver, basename: str, output_dir: str = "output", header_text: Optional[str] = None) -> Optional[str]:
    """生成带每页右上角追踪文字与右下页码的 PDF.

    利用 Chromium DevTools Page.printToPDF 的 headerTemplate/footerTemplate:
    - headerTemplate/footerTemplate 是 HTML 片段; 可使用 class 名称并靠 CSS 控制位置。
    - 占位符: <span class="pageNumber"></span> <span class="totalPages"></span>

    局限: header/footer 默认在纸张 margin 区域, 不是页面主体第一行; 若需强制正文下移已通过插入覆盖层实现。
    """
    try:
        if getattr(sys, 'frozen', False):
            base_dir = os.path.dirname(sys.executable)
        else:
            base_dir = os.getcwd()
        out_dir = os.path.join(base_dir, output_dir)
        os.makedirs(out_dir, exist_ok=True)
        # 构造 header/footer HTML
        hdr_html = ""
        if header_text:
            # 右上角: 使用 table 或 div + text-align:right
            hdr_html = (
                f"<div style='font-size:10px;width:100%;text-align:right;font-family:Microsoft YaHei,sans-serif;'>"
                f"{header_text}</div>"
            )
        ftr_html = (
            "<div style='font-size:10px;width:100%;text-align:right;font-family:Microsoft YaHei,sans-serif;'>"
            "页 <span class='pageNumber'></span> / <span class='totalPages'></span></div>"
        )
        params = {
            "landscape": False,
            "printBackground": True,
            "preferCSSPageSize": True,
            "displayHeaderFooter": True,
            "headerTemplate": hdr_html,
            "footerTemplate": ftr_html,
            # 适当留边, 避免覆盖正文: 单位英寸
            "marginTop": 0.6,
            "marginBottom": 0.6,
            "marginLeft": 0.4,
            "marginRight": 0.4,
        }
        pdf_b64 = driver.execute_cdp_cmd("Page.printToPDF", params)["data"]
        pdf_path = os.path.join(out_dir, f"{basename}.pdf")
        with open(pdf_path, "wb") as f:
            f.write(base64.b64decode(pdf_b64))
        return pdf_path
    except Exception as e:
        print(f"PDF 生成失败: {e}")
        return None


@dataclass
class ExcelContext:
    path: str
    sheet_name: str
    header_row_index: int  # 0-based (在原工作表中的行号)
    seq_col: int           # 序号列 index (0-based)
    waybill_col: int       # 物流单号列 index (0-based)
    data_rows: List[List[Optional[str]]]  # 表头之后的所有行

    def find_row_by_seq(self, seq_value: str) -> int:
        """返回在 data_rows 中的索引, 未找到返回 -1"""
        for i, row in enumerate(self.data_rows):
            if self.seq_col < len(row) and row[self.seq_col] is not None and str(row[self.seq_col]).strip() == seq_value.strip():
                return i
        return -1


def load_excel_sheet(path: str, sheet_name: str) -> ExcelContext:
    wb = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"工作表 {sheet_name} 不存在")
    ws = wb[sheet_name]
    # 读取所有行 (保留空行结构)
    raw_rows: List[List[Optional[str]]] = []
    for row in ws.iter_rows(values_only=True):
        raw_rows.append(list(row))
    if not raw_rows:
        raise ValueError("Excel 文件为空")

    # 在前 20 行内查找包含 '序号' 和 '物流单号' 的表头行
    header_row_index = -1
    seq_col = waybill_col = -1
    for i in range(min(20, len(raw_rows))):
        header_cells = [str(c).strip() if c is not None else "" for c in raw_rows[i]]
        if "序号" in header_cells and "物流单号" in header_cells:
            header_row_index = i
            seq_col = header_cells.index("序号")
            waybill_col = header_cells.index("物流单号")
            break
    if header_row_index == -1:
        raise ValueError("未找到 '序号' 或 '物流单号' 列, 请检查文件 (支持表头位于前20行)")

    data_rows = raw_rows[header_row_index + 1:]
    return ExcelContext(
        path=path,
        sheet_name=ws.title,
        header_row_index=header_row_index,
        seq_col=seq_col,
        waybill_col=waybill_col,
        data_rows=data_rows,
    )


class BatchUI:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("顺丰批量 PDF")
        self.root.attributes('-topmost', True)
        w, h = 480, 160
        sw, sh = self.root.winfo_screenwidth(), self.root.winfo_screenheight()
        margin = 8
        self.root.geometry(f"{w}x{h}+{sw-w-margin}+{margin}")

        # 状态数据
        self.excel_ctx: Optional[ExcelContext] = None
        self.order_no_var = tk.StringVar()
        self.current_row_index: Optional[int] = None  # 在 excel_ctx.rows (不含表头) 中的索引
        self.current_seq_value: Optional[str] = None  # 保存当前序号 (xu)
        self.driver: Optional[WebDriver] = None
        self.excel_path: Optional[str] = None
        self.sheet_btn_frame = None
        self.month_prefix: Optional[str] = None  # 从 sheet 名提取的首个数字序列 (X)

        # 第一行: 选择Excel
        top1 = tk.Frame(self.root)
        top1.pack(fill='x', pady=2)
        self.excel_label_var = tk.StringVar(value="未选择Excel")
        tk.Button(top1, text="选择Excel", width=12, command=self.choose_excel).pack(side='left', padx=4)
        tk.Label(top1, textvariable=self.excel_label_var, anchor='w').pack(side='left', padx=4)

        # 第二行: 输入序号 + 序号按钮
        top2 = tk.Frame(self.root)
        top2.pack(fill='x', pady=2)
        tk.Entry(top2, textvariable=self.order_no_var, width=12).pack(side='left', padx=4)
        tk.Button(top2, text="序号", width=8, command=self.set_order_no).pack(side='left', padx=4)
        self.seq_info_var = tk.StringVar(value="")
        tk.Label(top2, textvariable=self.seq_info_var, fg='#555').pack(side='left', padx=4)

        # 第三行: 控制按钮
        top3 = tk.Frame(self.root)
        top3.pack(fill='x', pady=4)
        self.btn_confirm = tk.Button(top3, text="确认", width=10, command=self.on_confirm, state=tk.DISABLED)
        self.btn_confirm.pack(side='left', padx=4)
        self.btn_next = tk.Button(top3, text="下一单", width=10, command=self.on_next, state=tk.DISABLED)
        self.btn_next.pack(side='left', padx=4)
        self.btn_end = tk.Button(top3, text="结束", width=10, command=self.on_end)
        self.btn_end.pack(side='left', padx=4)

        self.status_var = tk.StringVar(value="请选择 Excel, 输入序号, 点击 '序号'")
        tk.Label(self.root, textvariable=self.status_var, fg='#333').pack(fill='x', pady=4)

        self.root.protocol('WM_DELETE_WINDOW', self.on_end)

    # UI 事件
    def choose_excel(self):
        path = filedialog.askopenfilename(title="选择 Excel", filetypes=[("Excel", "*.xlsx")])
        if not path:
            return
        # 清除旧的 sheet 按钮
        if self.sheet_btn_frame:
            self.sheet_btn_frame.destroy()
            self.sheet_btn_frame = None
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            self.excel_path = path
            self.excel_label_var.set(os.path.basename(path))
            # 生成 sheet 按钮行
            self.sheet_btn_frame = tk.Frame(self.root)
            self.sheet_btn_frame.pack(fill='x', pady=2)
            for name in wb.sheetnames:
                b = tk.Button(self.sheet_btn_frame, text=name, width=10,
                              command=lambda n=name: self.load_sheet(n))
                b.pack(side='left', padx=2)
            self.status_var.set("请选择要使用的 sheet")
        except Exception as e:
            messagebox.showerror("错误", f"读取 Excel 失败: {e}")

    def load_sheet(self, sheet_name: str):
        if not self.excel_path:
            return
        try:
            ctx = load_excel_sheet(self.excel_path, sheet_name)
            self.excel_ctx = ctx
            self.current_row_index = None
            self.current_seq_value = None
            # 提取 sheet 名中的首个连续数字作为 X (如 '10月数据' -> '10')
            import re
            m = re.search(r'(\d+)', sheet_name)
            self.month_prefix = m.group(1) if m else None
            self.status_var.set(f"Sheet '{sheet_name}' 已加载, 输入序号后点击 '序号'")
            self.seq_info_var.set("")
        except Exception as e:
            messagebox.showerror("错误", f"加载 Sheet 失败: {e}")

    def set_order_no(self):
        if not self.excel_ctx:
            messagebox.showwarning("提示", "请先选择 Excel 文件")
            return
        val = self.order_no_var.get().strip()
        if not val:
            messagebox.showwarning("提示", "请输入序号数字")
            return
        idx = self.excel_ctx.find_row_by_seq(val)
        if idx == -1:
            messagebox.showwarning("提示", f"未找到序号 {val}")
            return
        self.current_row_index = idx
        self.current_seq_value = val  # 保存序号数字 xu
        waybill = self.get_current_waybill()
        # Excel 实际行号 = header_row_index(0-based) + 1 转 1-based + 1 数据偏移 + idx
        excel_row_num = self.excel_ctx.header_row_index + 2 + idx
        self.seq_info_var.set(f"行: {excel_row_num} 运单: {waybill}")
        self.status_var.set("准备打开网页, 请等待浏览器...")
        self.open_current_page()

    def get_current_waybill(self) -> Optional[str]:
        if self.excel_ctx is None or self.current_row_index is None:
            return None
        row = self.excel_ctx.data_rows[self.current_row_index]
        val = row[self.excel_ctx.waybill_col] if self.excel_ctx.waybill_col < len(row) else None
        return None if val is None else str(val).strip()

    def get_current_seq(self) -> Optional[str]:
        """从当前行读取序号列(防止 self.current_seq_value 丢失或为 None)."""
        if self.excel_ctx is None or self.current_row_index is None:
            return None
        row = self.excel_ctx.data_rows[self.current_row_index]
        if self.excel_ctx.seq_col >= len(row):
            return None
        cell = row[self.excel_ctx.seq_col]
        return None if cell is None else str(cell).strip()

    def open_current_page(self):
        waybill = self.get_current_waybill()
        if not waybill:
            messagebox.showwarning("提示", "该行物流单号为空")
            return
        if waybill == 'END':
            self.status_var.set("遇到 END, 程序结束")
            return
        # 启动浏览器 (若不存在)
        if self.driver is None:
            try:
                self.driver = create_driver()
            except Exception as e:
                messagebox.showerror("错误", f"创建浏览器失败: {e}")
                return
        url = BASE_URL.format(waybill=waybill)
        self.status_var.set(f"打开 {waybill} 中...")
        def _load():
            try:
                self.driver.get(url)
                self.status_var.set("请在浏览器中输入验证码并展开详情, 完成后点 '确认'")
                self.btn_confirm.config(state=tk.NORMAL)
            except Exception as e:
                self.status_var.set(f"页面加载失败: {e}")
        threading.Thread(target=_load, daemon=True).start()

    def on_confirm(self):
        waybill = self.get_current_waybill()
        if not waybill or not self.driver:
            return
        self.btn_confirm.config(state=tk.DISABLED)
        self.status_var.set("生成 PDF 中...")
        def _pdf():
            # 确保序号存在 (可能因切换/下一单后未重新赋值导致 None)
            if not self.current_seq_value:
                self.current_seq_value = self.get_current_seq() or "NA"
            # overlay 与文件名需要加 X月- 前缀: 若 month_prefix 存在则 'X月-' 否则空
            prefix = f"{self.month_prefix}月-" if self.month_prefix else ""
            overlay_text = f"{prefix}序号{self.current_seq_value}-{waybill}"
            try:
                js = (
                    "(function(){"  # 创建顶部行并推下内容
                    "var id='__sf_overlay__';" 
                    "var old=document.getElementById(id);" 
                    "if(old){old.remove();}" 
                    # 创建容器 wrapper，如果没有则插入 body 最前面
                    "var body=document.body;" 
                    "var div=document.createElement('div');" 
                    "div.id=id;" 
                    "div.textContent='" + overlay_text + "';" 
                    # 样式: 正常文档流顶部一行，右对齐，白底黑字
                    "div.style.position='relative';" 
                    "div.style.width='100%';" 
                    "div.style.boxSizing='border-box';" 
                    "div.style.textAlign='right';" 
                    "div.style.font='16px/1.4 \"Microsoft YaHei\",sans-serif';" 
                    "div.style.color='#000';" 
                    "div.style.background='#fff';" 
                    "div.style.padding='6px 12px';" 
                    "div.style.borderBottom='1px solid #ddd';" 
                    "div.style.margin='0';" 
                    "if(body.firstChild){body.insertBefore(div, body.firstChild);}else{body.appendChild(div);}" 
                    "})();"
                )
                self.driver.execute_script(js)
            except Exception as e:
                print(f"注入覆盖文字失败: {e}")

            # 使用自定义文件名 序号{xu}-{waybill}.pdf
            custom_name = f"{prefix}序号{self.current_seq_value}-{waybill}"
            # header_text 与 overlay_text 保持一致 (追踪文字段 BB)
            pdf_path = print_to_pdf(self.driver, custom_name, header_text=overlay_text)
            if pdf_path:
                self.status_var.set(f"PDF 已生成: {os.path.basename(pdf_path)} 点击 '下一单'")
                self.btn_next.config(state=tk.NORMAL)
            else:
                self.status_var.set("生成失败, 可重试 '确认'")
                self.btn_confirm.config(state=tk.NORMAL)
        threading.Thread(target=_pdf, daemon=True).start()

    def on_next(self):
        if self.excel_ctx is None or self.current_row_index is None:
            return
        # 关闭当前页面? 不关闭浏览器, 直接继续
        self.btn_next.config(state=tk.DISABLED)
        self.status_var.set("读取下一行...")
        self.current_row_index += 1
        # 立即根据新行刷新序号缓存，避免出现 None
        self.current_seq_value = self.get_current_seq()
        if self.current_row_index >= len(self.excel_ctx.data_rows):
            self.status_var.set("已到文件末尾, 程序结束")
            return
        waybill = self.get_current_waybill()
        if waybill == 'END':
            self.status_var.set("遇到 END, 程序结束")
            return
        excel_row_num = self.excel_ctx.header_row_index + 2 + self.current_row_index
        seq_display = self.current_seq_value or "?"
        self.seq_info_var.set(f"行: {excel_row_num} 序号: {seq_display} 运单: {waybill}")
        self.btn_confirm.config(state=tk.DISABLED)
        self.open_current_page()

    def on_end(self):
        try:
            if self.driver:
                self.driver.quit()
        finally:
            self.root.destroy()

    def run(self):
        self.root.mainloop()


def main():
    ui = BatchUI()
    ui.run()
    return 0


if __name__ == '__main__':
    raise SystemExit(main())
